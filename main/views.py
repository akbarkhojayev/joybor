from rest_framework import generics, status, filters, mixins
from rest_framework.exceptions import NotFound
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from .models import *
from .serializers import *
from .permisssions import *
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.exceptions import PermissionDenied
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi


# ==================== AUTH VIEWS ====================
class RegisterView(generics.CreateAPIView):
    serializer_class = RegisterSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(
            {"message": "Foydalanuvchi muvaffaqiyatli ro'yxatdan o'tdi"},
            status=status.HTTP_201_CREATED
        )


class GoogleLoginView(APIView):
    """Google OAuth2 - login yoki register (avtomatik)"""
    permission_classes = [AllowAny]
    serializer_class = GoogleTokenSerializer

    @swagger_auto_schema(request_body=GoogleTokenSerializer)
    def post(self, request):
        serializer = GoogleTokenSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        token = serializer.validated_data['token']
        google_user = self._verify_google_token(token)
        if not google_user:
            return Response({"error": "Noto'g'ri yoki muddati o'tgan Google token"}, status=status.HTTP_400_BAD_REQUEST)

        email = google_user.get('email')
        if not email:
            return Response({"error": "Google akkauntda email topilmadi"}, status=status.HTTP_400_BAD_REQUEST)

        from rest_framework_simplejwt.tokens import RefreshToken

        # Mavjud userni topish
        try:
            user = User.objects.get(email=email)
            is_new = False
        except User.DoesNotExist:
            # Yangi student yaratish
            username = self._generate_username(email)
            user = User.objects.create(
                email=email,
                username=username,
                first_name=google_user.get('given_name', ''),
                last_name=google_user.get('family_name', ''),
                role='student',
            )
            user.set_unusable_password()
            user.save()
            UserProfile.objects.create(user=user)
            is_new = True

        refresh = RefreshToken.for_user(user)
        return Response({
            "is_new_user": is_new,
            "message": "Ro'yxatdan o'tdingiz" if is_new else "Muvaffaqiyatli kirish",
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "role": user.role,
            }
        }, status=status.HTTP_201_CREATED if is_new else status.HTTP_200_OK)

    def _verify_google_token(self, token):
        import requests as req
        from django.conf import settings

        try:
            resp = req.get('https://oauth2.googleapis.com/tokeninfo', params={'id_token': token}, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                client_id = getattr(settings, 'GOOGLE_CLIENT_ID', '')
                if client_id and data.get('aud') != client_id:
                    return None
                return data
        except Exception:
            pass

        try:
            resp = req.get('https://www.googleapis.com/oauth2/v3/userinfo',
                           headers={'Authorization': f'Bearer {token}'}, timeout=10)
            if resp.status_code == 200:
                return resp.json()
        except Exception:
            pass

        return None

    def _generate_username(self, email):
        import re
        base = re.sub(r'[^a-zA-Z0-9]', '_', email.split('@')[0])
        username = base
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"{base}_{counter}"
            counter += 1
        return username


class GoogleRegisterView(APIView):
    """Google OAuth2 orqali faqat REGISTER (faqat student uchun)"""
    permission_classes = [AllowAny]
    serializer_class = GoogleTokenSerializer  # Swagger uchun

    @swagger_auto_schema(request_body=GoogleTokenSerializer)
    def post(self, request):
        serializer = GoogleTokenSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        token = serializer.validated_data['token']

        google_user = self._verify_google_token(token)
        if not google_user:
            return Response({"error": "Noto'g'ri yoki muddati o'tgan Google token"}, status=status.HTTP_400_BAD_REQUEST)

        email = google_user.get('email')
        if not email:
            return Response({"error": "Google akkauntda email topilmadi"}, status=status.HTTP_400_BAD_REQUEST)

        # Allaqachon ro'yxatdan o'tganmi tekshirish
        if User.objects.filter(email=email).exists():
            return Response({
                "error": "Bu Google akkaunt allaqachon ro'yxatdan o'tgan. Login qiling.",
                "email": email,
                "is_registered": True,
            }, status=status.HTTP_400_BAD_REQUEST)

        # Yangi student yaratish
        username = self._generate_username(email)
        user = User.objects.create(
            email=email,
            username=username,
            first_name=google_user.get('given_name', ''),
            last_name=google_user.get('family_name', ''),
            role='student',
        )
        user.set_unusable_password()
        user.save()

        UserProfile.objects.create(user=user)

        from rest_framework_simplejwt.tokens import RefreshToken
        refresh = RefreshToken.for_user(user)

        return Response({
            "message": "Muvaffaqiyatli ro'yxatdan o'tish",
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "role": user.role,
            }
        }, status=status.HTTP_201_CREATED)

    def _verify_google_token(self, token):
        import requests as req
        from django.conf import settings

        # ID token tekshirish
        try:
            resp = req.get(
                'https://oauth2.googleapis.com/tokeninfo',
                params={'id_token': token},
                timeout=10
            )
            if resp.status_code == 200:
                data = resp.json()
                client_id = getattr(settings, 'GOOGLE_CLIENT_ID', '')
                if client_id and data.get('aud') != client_id:
                    return None
                return data
        except Exception:
            pass

        # Access token tekshirish (fallback)
        try:
            resp = req.get(
                'https://www.googleapis.com/oauth2/v3/userinfo',
                headers={'Authorization': f'Bearer {token}'},
                timeout=10
            )
            if resp.status_code == 200:
                return resp.json()
        except Exception:
            pass

        return None

    def _generate_username(self, email):
        import re
        base = re.sub(r'[^a-zA-Z0-9]', '_', email.split('@')[0])
        username = base
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"{base}_{counter}"
            counter += 1
        return username


class CheckUsernameView(APIView):
    """Username mavjudligini tekshirish - GET yoki POST"""
    permission_classes = [AllowAny]

    def _check(self, username):
        if not username:
            return None, Response(
                {"error": "username kiritilmagan"},
                status=status.HTTP_400_BAD_REQUEST
            )
        exists = User.objects.filter(username__iexact=username).exists()
        return Response({
            "username": username,
            "available": not exists,
            "message": "Bu username band" if exists else "Bu username bo'sh"
        }), None

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter('username', openapi.IN_QUERY, type=openapi.TYPE_STRING, description="Tekshiriladigan username")
        ]
    )
    def get(self, request):
        username = request.query_params.get('username', '').strip()
        result, err = self._check(username)
        return err or result

    @swagger_auto_schema(request_body=CheckUsernameSerializer)
    def post(self, request):
        username = request.data.get('username', '').strip()
        result, err = self._check(username)
        return err or result


class CurrentUserView(generics.RetrieveUpdateAPIView):
    """Joriy foydalanuvchi ma'lumotlarini ko'rish va tahrirlash"""
    serializer_class = UserMeSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        user = self.request.user

        # SUPERUSER ⇒ 404 bo'lmaydi
        if user.is_superuser:
            # Agar profil mavjud bo'lsa — qaytaramiz
            try:
                return user.profile
            except UserProfile.DoesNotExist:
                # Profil yo'q bo'lsa yaratamiz
                return UserProfile.objects.create(user=user)

        # ODDIY USER ⇒ Profil bo'lmasa yaratamiz
        try:
            return user.profile
        except UserProfile.DoesNotExist:
            return UserProfile.objects.create(user=user)


# ==================== USER VIEWS ====================
class UserListView(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['username', 'email']
    filterset_fields = ['role', 'is_active']


class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]


class UserCreateView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]


# ==================== USER PROFILE VIEWS ====================
class UserProfileListView(generics.ListAPIView):
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return UserProfile.objects.none()
        if self.request.user.is_superuser:
            return UserProfile.objects.all()
        return UserProfile.objects.filter(user=self.request.user)


class UserProfileDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return UserProfile.objects.none()
        if self.request.user.is_superuser:
            return UserProfile.objects.all()
        return UserProfile.objects.filter(user=self.request.user)


# ==================== PROVINCE VIEWS ====================
class ProvinceListView(generics.ListCreateAPIView):
    queryset = Province.objects.all()
    serializer_class = ProvinceSerializer
    permission_classes = [AllowAny]


class ProvinceDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Province.objects.all()
    serializer_class = ProvinceSerializer
    permission_classes = [AllowAny]


# ==================== DISTRICT VIEWS ====================
class DistrictListView(generics.ListCreateAPIView):
    queryset = District.objects.all()
    serializer_class = DistrictSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['province']


class DistrictDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = District.objects.all()
    serializer_class = DistrictSerializer
    permission_classes = [AllowAny]


# ==================== UNIVERSITY VIEWS ====================
class UniversityListView(generics.ListCreateAPIView):
    queryset = University.objects.all()
    serializer_class = UniversitySerializer
    permission_classes = [AllowAny]


class UniversityDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = University.objects.all()
    serializer_class = UniversitySerializer
    permission_classes = [AllowAny]


# ==================== AMENITY VIEWS ====================
class AmenityListView(generics.ListCreateAPIView):
    queryset = Amenity.objects.all()
    serializer_class = AmenitySerializer
    permission_classes = [IsAdminOrDormitoryAdmin]


class AmenityDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Amenity.objects.all()
    serializer_class = AmenitySerializer
    permission_classes = [IsAdminOrDormitoryAdmin]

# ==================== DORMITORY VIEWS ====================
class DormitoryListView(generics.ListAPIView):
    serializer_class = DormitorySerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['name', 'address']
    filterset_fields = ['university', 'is_active']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Dormitory.objects.filter(is_active=True)
        if self.request.user.is_authenticated and hasattr(self.request.user, 'role') and self.request.user.role == 'admin':
            return Dormitory.objects.filter(admin=self.request.user)
        return Dormitory.objects.filter(is_active=True)


class DormitoryCreateView(generics.CreateAPIView):
    queryset = Dormitory.objects.all()
    serializer_class = DormitorySerializer
    permission_classes = [IsAdminUser]


class DormitoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Dormitory.objects.all()
    serializer_class = DormitorySerializer
    permission_classes = [IsOwnerOrIsAdmin]

class DormitoryImageListView(generics.ListCreateAPIView):
    serializer_class = DormitoryImageSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [DjangoFilterBackend]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        user = self.request.user

        # Superuser hammasini ko‘ra oladi
        if user.is_superuser:
            return DormitoryImage.objects.all()

        # Admin faqat o‘z dormitory rasmlarini
        return DormitoryImage.objects.filter(
            dormitory__admin=user
        )

    def perform_create(self, serializer):
        user = self.request.user

        # user admin bo‘lgan dormitoryni topamiz
        dormitory = Dormitory.objects.filter(admin=user).first()
        if not dormitory:
            raise PermissionDenied(
                "Sizga tegishli dormitory topilmadi"
            )

        serializer.save(dormitory=dormitory)


class DormitoryImageDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DormitoryImageSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]

    def get_queryset(self):
        user = self.request.user

        if user.is_superuser:
            return DormitoryImage.objects.all()

        return DormitoryImage.objects.filter(
            dormitory__admin=user
        )



# ==================== RULE VIEWS ====================
class RuleListView(generics.ListCreateAPIView):
    queryset = Rule.objects.all()
    serializer_class = RuleSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['dormitory']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Rule.objects.none()
        
        user = self.request.user
        if user.is_superuser:
            return Rule.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Rule.objects.filter(dormitory__admin=user)
        # Barcha foydalanuvchilar dormitory qoidalarini ko'rishi mumkin
        return Rule.objects.all()

    def perform_create(self, serializer):
        user = self.request.user
        if not hasattr(user, 'role') or user.role != 'admin':
            raise PermissionDenied("Faqat admin qoida qo'sha oladi")
        
        try:
            dormitory = Dormitory.objects.get(admin=user)
        except Dormitory.DoesNotExist:
            raise PermissionDenied("Sizga tegishli yotoqxona topilmadi")
        
        serializer.save(dormitory=dormitory)



class RuleDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = RuleSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Rule.objects.none()
        
        user = self.request.user
        if user.is_superuser:
            return Rule.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            # Admin faqat o'z yotoqxonasining qoidalarini ko'radi
            return Rule.objects.filter(dormitory__admin=user)
        return Rule.objects.none()


# ==================== FLOOR VIEWS ====================
class FloorListView(generics.ListCreateAPIView):
    serializer_class = FloorSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['gender']
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Floor.objects.none()
        
        user = self.request.user
        if user.is_superuser:
            return Floor.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            # Admin faqat o'z yotoqxonasining qavatlarini ko'radi
            return Floor.objects.filter(dormitory__admin=user)
        return Floor.objects.none()

    def perform_create(self, serializer):
        user = self.request.user

        if user.role != 'admin':
            raise PermissionDenied("Siz yotoqxona admini emassiz !!!")

        try:
            dormitory = Dormitory.objects.get(admin=user)
        except Dormitory.DoesNotExist:
            raise PermissionDenied("Sizga tegishli yotoqxona mavjud emas !!!")

        serializer.save(dormitory=dormitory)

class FloorDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = FloorSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Floor.objects.none()
        
        user = self.request.user
        if user.is_superuser:
            return Floor.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            # Admin faqat o'z yotoqxonasining qavatlarini ko'radi
            return Floor.objects.filter(dormitory__admin=user)
        return Floor.objects.none()


# ==================== ROOM VIEWS ====================
class RoomListView(generics.ListCreateAPIView):
    serializer_class = RoomSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['floor', 'status', 'gender']
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Room.objects.none()
        
        user = self.request.user
        if user.is_superuser:
            return Room.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            # Admin faqat o'z yotoqxonasining xonalarini ko'radi
            return Room.objects.filter(floor__dormitory__admin=user)
        return Room.objects.none()


class RoomDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = RoomSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Room.objects.none()
        
        user = self.request.user
        if user.is_superuser:
            return Room.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            # Admin faqat o'z yotoqxonasining xonalarini ko'radi
            return Room.objects.filter(floor__dormitory__admin=user)
        return Room.objects.none()


# ==================== STUDENT VIEWS ====================
class StudentListView(generics.ListAPIView):
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]  # Hamma authenticated user ko'ra oladi
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['name', 'last_name', 'passport', 'phone']
    filterset_fields = ['dormitory', 'floor', 'room', 'course', 'gender', 'status']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Student.objects.none()
        user = self.request.user
        
        # Superuser - barcha talabalarni ko'radi
        if user.is_superuser:
            return Student.objects.all()
        
        # Admin - o'z yotoqxonasidagi talabalarni ko'radi
        elif hasattr(user, 'role') and user.role == 'admin':
            return Student.objects.filter(dormitory__admin=user)
        
        # Sardor (Floor Leader) - faqat o'z qavatidagi talabalarni ko'radi
        elif hasattr(user, 'role') and user.role == 'sardor':
            try:
                floor_leader = FloorLeader.objects.get(user=user)
                # Faqat o'z qavatidagi talabalar
                return Student.objects.filter(
                    floor=floor_leader.floor,
                    dormitory=floor_leader.floor.dormitory,
                    is_active=True  # Faqat faol talabalar
                )
            except FloorLeader.DoesNotExist:
                return Student.objects.none()
        
        # Student - faqat o'zini ko'radi
        return Student.objects.filter(user=user)


class StudentCreateView(generics.CreateAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]


class StudentDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]  # Hamma authenticated user ko'ra oladi

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Student.objects.none()
        user = self.request.user
        
        # Superuser - barcha talabalarni ko'radi
        if user.is_superuser:
            return Student.objects.all()
        
        # Admin - o'z yotoqxonasidagi talabalarni ko'radi va tahrirlashi mumkin
        elif hasattr(user, 'role') and user.role == 'admin':
            return Student.objects.filter(dormitory__admin=user)
        
        # Sardor (Floor Leader) - faqat o'z qavatidagi talabalarni ko'radi (faqat o'qish)
        elif hasattr(user, 'role') and user.role == 'sardor':
            try:
                floor_leader = FloorLeader.objects.get(user=user)
                return Student.objects.filter(
                    floor=floor_leader.floor,
                    dormitory=floor_leader.floor.dormitory,
                    is_active=True
                )
            except FloorLeader.DoesNotExist:
                return Student.objects.none()
        
        # Student - faqat o'zini ko'radi
        return Student.objects.filter(user=user)
    
    def update(self, request, *args, **kwargs):
        """Faqat admin tahrirlashi mumkin"""
        user = request.user
        if not user.is_superuser and (not hasattr(user, 'role') or user.role != 'admin'):
            return Response(
                {"error": "Faqat admin talaba ma'lumotlarini tahrirlashi mumkin"},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        """Faqat admin o'chirishi mumkin"""
        user = request.user
        if not user.is_superuser and (not hasattr(user, 'role') or user.role != 'admin'):
            return Response(
                {"error": "Faqat admin talabani o'chirishi mumkin"},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)


# ==================== APPLICATION VIEWS ====================
class ApplicationListView(generics.ListAPIView):
    serializer_class = ApplicationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['name', 'last_name', 'passport']
    filterset_fields = ['dormitory', 'status']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Application.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Application.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Application.objects.filter(dormitory__admin=user)
        return Application.objects.filter(user=user)


class ApplicationCreateView(generics.CreateAPIView):
    queryset = Application.objects.all()
    serializer_class = ApplicationSerializer
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [AllowAny]
    
    def perform_create(self, serializer):
        # Agar foydalanuvchi tizimga kirgan bo'lsa, user ni biriktirish
        if self.request.user.is_authenticated:
            serializer.save(user=self.request.user)
        else:
            serializer.save()


class ApplicationDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ApplicationSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Application.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Application.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Application.objects.filter(dormitory__admin=user)
        return Application.objects.filter(user=user)

class ApplicationApproveView(generics.UpdateAPIView):
    """Arizani tasdiqlash va avtomatik Student yaratish"""
    queryset = Application.objects.all()
    serializer_class = ApplicationApproveSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Application.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Application.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Application.objects.filter(dormitory__admin=user)
        return Application.objects.none()
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        return Response({
            "message": "Ariza tasdiqlandi va talaba ro'yxatga olindi",
            "admin_comment": instance.admin_comment,
            "application": ApplicationSerializer(instance).data
        }, status=status.HTTP_200_OK)


class ApplicationRejectView(generics.UpdateAPIView):
    """Arizani rad etish"""
    queryset = Application.objects.all()
    serializer_class = ApplicationRejectSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Application.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Application.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Application.objects.filter(dormitory__admin=user)
        return Application.objects.none()
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        return Response({
            "message": "Ariza rad etildi",
            "admin_comment": instance.admin_comment,
            "application": ApplicationSerializer(instance).data
        }, status=status.HTTP_200_OK)

# ==================== PAYMENT VIEWS ====================
class PaymentListView(generics.ListAPIView):
    serializer_class = PaymentSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['student', 'status', 'method', 'dormitory']
    ordering_fields = ['paid_date', 'amount', 'valid_until']
    ordering = ['-paid_date']  # Eng yangi to'lovlar birinchi

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Payment.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Payment.objects.select_related('student', 'dormitory').all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Payment.objects.select_related('student', 'dormitory').filter(dormitory__admin=user)
        return Payment.objects.none()

class PaymentCreateView(generics.CreateAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]

    def perform_create(self, serializer):
        student = serializer.validated_data['student']
        
        # Dormitory ni aniqlash
        dormitory = None
        
        # 1. Agar student floor ga biriktirilgan bo'lsa
        if student.floor and student.floor.dormitory:
            dormitory = student.floor.dormitory
        # 2. Agar student dormitory ga biriktirilgan bo'lsa (lekin floor yo'q)
        elif student.dormitory:
            dormitory = student.dormitory
        # 3. Agar admin o'z yotoqxonasiga to'lov qo'shayotgan bo'lsa
        elif hasattr(self.request.user, 'role') and self.request.user.role == 'admin':
            dormitory = Dormitory.objects.filter(admin=self.request.user).first()
        
        if not dormitory:
            raise serializers.ValidationError({
                "error": "Talaba hech qaysi yotoqxonaga biriktirilmagan. Avval talabani yotoqxonaga biriktiring."
            })
        
        # To'lovni saqlash
        payment = serializer.save(dormitory=dormitory)
        
        # Talabaga bildirishnoma yuborish
        if student.user and payment.status == 'APPROVED':
            ApplicationNotification.objects.create(
                user=student.user,
                message=f"To'lovingiz qabul qilindi. Summa: {payment.amount} so'm. Amal qilish muddati: {payment.valid_until or 'Korsatilmagan'}"
            )

class PaymentDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = PaymentSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Payment.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Payment.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Payment.objects.filter(dormitory__admin=user)
        return Payment.objects.none()
    
    def perform_update(self, serializer):
        """To'lov yangilanganda talabaga bildirishnoma yuborish"""
        old_status = self.get_object().status
        payment = serializer.save()
        
        # Agar status APPROVED ga o'zgartirilgan bo'lsa
        if old_status != 'APPROVED' and payment.status == 'APPROVED':
            if payment.student.user:
                ApplicationNotification.objects.create(
                    user=payment.student.user,
                    message=f"To'lovingiz tasdiqlandi. Summa: {payment.amount} so'm. Amal qilish muddati: {payment.valid_until or 'Ko\'rsatilmagan'}"
                )
        
        # Agar status CANCELLED ga o'zgartirilgan bo'lsa
        elif old_status != 'CANCELLED' and payment.status == 'CANCELLED':
            if payment.student.user:
                ApplicationNotification.objects.create(
                    user=payment.student.user,
                    message=f"To'lovingiz bekor qilindi. Summa: {payment.amount} so'm. Sabab: {payment.comment or 'Ko\'rsatilmagan'}"
                )


# ==================== TASK VIEWS ====================
class TaskListView(generics.ListAPIView):
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Task.objects.none()
        return Task.objects.filter(user=self.request.user)


class TaskCreateView(generics.CreateAPIView):
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class TaskDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Task.objects.none()
        return Task.objects.filter(user=self.request.user)


# ==================== APARTMENT VIEWS ====================
class ApartmentListView(generics.ListAPIView):
    serializer_class = ApartmentSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['title', 'exact_address']
    filterset_fields = ['province', 'room_type', 'gender', 'is_active']

    def get_queryset(self):
        return Apartment.objects.filter(is_active=True)


class ApartmentCreateView(generics.CreateAPIView):
    serializer_class = ApartmentSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ApartmentDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ApartmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Apartment.objects.filter(is_active=True)
        if self.request.method == 'GET':
            return Apartment.objects.filter(is_active=True)
        return Apartment.objects.filter(user=self.request.user)


# ==================== APARTMENT IMAGE VIEWS ====================
class ApartmentImageListView(generics.ListCreateAPIView):
    queryset = ApartmentImage.objects.all()
    serializer_class = ApartmentImageSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['apartment']


class ApartmentImageDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ApartmentImage.objects.all()
    serializer_class = ApartmentImageSerializer
    permission_classes = [IsAuthenticated]

# ==================== FLOOR LEADER VIEWS ====================
class FloorLeaderListView(generics.ListCreateAPIView):
    serializer_class = FloorLeaderSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['floor']
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return FloorLeader.objects.none()
        
        user = self.request.user
        if user.is_superuser:
            return FloorLeader.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            # Admin faqat o'z yotoqxonasining qavat sardorlarini ko'radi
            return FloorLeader.objects.filter(floor__dormitory__admin=user)
        return FloorLeader.objects.none()


class FloorLeaderDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = FloorLeaderSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return FloorLeader.objects.none()
        
        user = self.request.user
        if user.is_superuser:
            return FloorLeader.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            # Admin faqat o'z yotoqxonasining qavat sardorlarini ko'radi
            return FloorLeader.objects.filter(floor__dormitory__admin=user)
        return FloorLeader.objects.none()


# ==================== ATTENDANCE SESSION VIEWS ====================
class AttendanceSessionListView(generics.ListAPIView):
    serializer_class = AttendanceSessionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['floor', 'date']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return AttendanceSession.objects.none()
        user = self.request.user
        if user.is_superuser:
            return AttendanceSession.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return AttendanceSession.objects.filter(floor__dormitory__admin=user)
        elif hasattr(user, 'role') and user.role == 'sardor':
            floor_leader = FloorLeader.objects.filter(user=user).first()
            if floor_leader:
                return AttendanceSession.objects.filter(floor=floor_leader.floor)
        return AttendanceSession.objects.none()


class AttendanceSessionCreateView(generics.CreateAPIView):
    queryset = AttendanceSession.objects.all()
    serializer_class = AttendanceSessionSerializer
    permission_classes = [IsAuthenticated]


class AttendanceSessionDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = AttendanceSessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return AttendanceSession.objects.none()
        user = self.request.user
        if user.is_superuser:
            return AttendanceSession.objects.all()
        elif hasattr(user, 'role') and user.role == 'sardor':
            floor_leader = FloorLeader.objects.filter(user=user).first()
            if floor_leader:
                return AttendanceSession.objects.filter(floor=floor_leader.floor)
        return AttendanceSession.objects.none()


# ==================== ATTENDANCE RECORD VIEWS ====================
class AttendanceRecordListView(generics.ListCreateAPIView):
    queryset = AttendanceRecord.objects.all()
    serializer_class = AttendanceRecordSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['session', 'student', 'status']


class AttendanceRecordDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = AttendanceRecord.objects.all()
    serializer_class = AttendanceRecordSerializer
    permission_classes = [IsAuthenticated]


# ==================== COLLECTION VIEWS ====================
class CollectionListView(generics.ListAPIView):
    serializer_class = CollectionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['floor']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Collection.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Collection.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Collection.objects.filter(floor__dormitory__admin=user)
        elif hasattr(user, 'role') and user.role == 'sardor':
            floor_leader = FloorLeader.objects.filter(user=user).first()
            if floor_leader:
                return Collection.objects.filter(floor=floor_leader.floor)
        return Collection.objects.none()


class CollectionCreateView(generics.CreateAPIView):
    queryset = Collection.objects.all()
    serializer_class = CollectionSerializer
    permission_classes = [IsFloorLeader]


class CollectionDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CollectionSerializer
    permission_classes = [IsFloorLeader]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Collection.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Collection.objects.all()
        elif hasattr(user, 'role') and user.role == 'sardor':
            floor_leader = FloorLeader.objects.filter(user=user).first()
            if floor_leader:
                return Collection.objects.filter(floor=floor_leader.floor)
        return Collection.objects.none()


# ==================== COLLECTION RECORD VIEWS ====================
class CollectionRecordListView(generics.ListCreateAPIView):
    queryset = CollectionRecord.objects.all()
    serializer_class = CollectionRecordSerializer
    permission_classes = [IsFloorLeader]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['collection', 'student', 'status']


class CollectionRecordDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CollectionRecord.objects.all()
    serializer_class = CollectionRecordSerializer
    permission_classes = [IsFloorLeader]


# ==================== TASK FOR LEADER VIEWS ====================
class TaskForLeaderListView(generics.ListAPIView):
    serializer_class = TaskForLeaderSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return TaskForLeader.objects.none()
        user = self.request.user
        if hasattr(user, 'role') and user.role == 'sardor':
            return TaskForLeader.objects.filter(user=user)
        elif hasattr(user, 'role') and user.role == 'admin':
            floor_leaders = FloorLeader.objects.filter(floor__dormitory__admin=user)
            return TaskForLeader.objects.filter(user__in=[fl.user for fl in floor_leaders])
        return TaskForLeader.objects.none()


class TaskForLeaderCreateView(generics.CreateAPIView):
    queryset = TaskForLeader.objects.all()
    serializer_class = TaskForLeaderSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]


class TaskForLeaderDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = TaskForLeaderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return TaskForLeader.objects.none()
        user = self.request.user
        if hasattr(user, 'role') and user.role == 'sardor':
            return TaskForLeader.objects.filter(user=user)
        elif hasattr(user, 'role') and user.role == 'admin':
            floor_leaders = FloorLeader.objects.filter(floor__dormitory__admin=user)
            return TaskForLeader.objects.filter(user__in=[fl.user for fl in floor_leaders])
        return TaskForLeader.objects.none()


# ==================== DUTY SCHEDULE VIEWS ====================
class DutyScheduleListView(generics.ListCreateAPIView):
    queryset = DutySchedule.objects.all()
    serializer_class = DutyScheduleSerializer
    permission_classes = [IsFloorLeader]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['floor', 'room', 'date']


class DutyScheduleDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = DutySchedule.objects.all()
    serializer_class = DutyScheduleSerializer
    permission_classes = [IsFloorLeader]


# ==================== STUDENT DASHBOARD VIEWS ====================
class StudentDashboardView(APIView):
    """Talaba uchun to'liq dashboard ma'lumotlari"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            student = Student.objects.select_related(
                'user', 'dormitory', 'floor', 'room', 'province', 'district'
            ).get(user=request.user)
            
            serializer = StudentDashboardSerializer(student)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Student.DoesNotExist:
            return Response(
                {"error": "Siz hali talaba sifatida ro'yxatdan o'tmagansiz"}, 
                status=status.HTTP_404_NOT_FOUND
            )


class StudentPaymentsView(generics.ListAPIView):
    """Talabaning barcha to'lovlari"""
    serializer_class = PaymentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'method']
    
    def get_queryset(self):
        try:
            student = Student.objects.get(user=self.request.user)
            return Payment.objects.filter(student=student).order_by('-paid_date')
        except Student.DoesNotExist:
            return Payment.objects.none()


class StudentRoommatesView(generics.ListAPIView):
    """Talabaning xonadoshlari"""
    serializer_class = RoommateSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        try:
            student = Student.objects.get(user=self.request.user)
            if student.room:
                return Student.objects.filter(room=student.room).exclude(id=student.id)
            return Student.objects.none()
        except Student.DoesNotExist:
            return Student.objects.none()


class StudentAttendanceView(generics.ListAPIView):
    """Talabaning davomat ma'lumotlari"""
    serializer_class = AttendanceRecordSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'session__date']
    
    def get_queryset(self):
        try:
            student = Student.objects.get(user=self.request.user)
            return AttendanceRecord.objects.filter(student=student).order_by('-created_at')
        except Student.DoesNotExist:
            return AttendanceRecord.objects.none()


class StudentCollectionsView(generics.ListAPIView):
    """Talabaning yig'im ma'lumotlari"""
    serializer_class = CollectionRecordSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'collection']
    
    def get_queryset(self):
        try:
            student = Student.objects.get(user=self.request.user)
            return CollectionRecord.objects.filter(student=student).select_related('collection')
        except Student.DoesNotExist:
            return CollectionRecord.objects.none()


class NotificationsView(APIView):
    """Barcha bildirishnomalar - bitta unified endpoint (UserNotification + ApplicationNotification)"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        if getattr(self, 'swagger_fake_view', False):
            return Response([])
        
        # UserNotification larni olish
        user_notifications = UserNotification.objects.filter(
            user=request.user
        ).select_related('notification').order_by('-received_at')
        
        # ApplicationNotification larni olish
        app_notifications = ApplicationNotification.objects.filter(
            user=request.user
        ).order_by('-created_at')
        
        # Ikkalasini birlashtirib, vaqt bo'yicha saralash
        all_notifications = list(user_notifications) + list(app_notifications)
        all_notifications.sort(
            key=lambda x: x.received_at if isinstance(x, UserNotification) else x.created_at,
            reverse=True
        )
        
        # Filter by is_read if provided
        is_read_filter = request.query_params.get('is_read')
        if is_read_filter is not None:
            is_read_bool = is_read_filter.lower() in ['true', '1', 'yes']
            all_notifications = [n for n in all_notifications if n.is_read == is_read_bool]
        
        # Pagination
        from rest_framework.pagination import PageNumberPagination
        paginator = PageNumberPagination()
        paginator.page_size = 20
        
        page = paginator.paginate_queryset(all_notifications, request)
        if page is not None:
            serializer = UnifiedNotificationSerializer(page, many=True)
            return paginator.get_paginated_response(serializer.data)
        
        serializer = UnifiedNotificationSerializer(all_notifications, many=True)
        return Response(serializer.data)



class MarkNotificationAsReadView(APIView):
    """Bildirishnomani o'qilgan deb belgilash (har ikki tur uchun)"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        notification_type = request.data.get('type')  # 'user' yoki 'application'
        notification_id = request.data.get('id')
        
        if not notification_type or not notification_id:
            return Response(
                {"error": "type va id majburiy"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            if notification_type == 'user':
                notification = UserNotification.objects.get(
                    id=notification_id,
                    user=request.user
                )
                notification.is_read = True
                notification.save()
                return Response(
                    {"message": "Bildirishnoma o'qilgan deb belgilandi"},
                    status=status.HTTP_200_OK
                )
            
            elif notification_type == 'application':
                notification = ApplicationNotification.objects.get(
                    id=notification_id,
                    user=request.user
                )
                notification.is_read = True
                notification.save()
                return Response(
                    {"message": "Bildirishnoma o'qilgan deb belgilandi"},
                    status=status.HTTP_200_OK
                )
            
            else:
                return Response(
                    {"error": "Noto'g'ri type. 'user' yoki 'application' bo'lishi kerak"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        except (UserNotification.DoesNotExist, ApplicationNotification.DoesNotExist):
            return Response(
                {"error": "Bildirishnoma topilmadi"},
                status=status.HTTP_404_NOT_FOUND
            )


class MarkAllNotificationsAsReadView(APIView):
    """Barcha bildirishnomalarni o'qilgan deb belgilash"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # UserNotification larni yangilash
        user_count = UserNotification.objects.filter(
            user=request.user,
            is_read=False
        ).update(is_read=True)
        
        # ApplicationNotification larni yangilash
        app_count = ApplicationNotification.objects.filter(
            user=request.user,
            is_read=False
        ).update(is_read=True)
        
        total = user_count + app_count
        
        return Response(
            {"message": f"{total} ta bildirishnoma o'qilgan deb belgilandi"},
            status=status.HTTP_200_OK
        )


class UnreadNotificationCountView(APIView):
    """O'qilmagan bildirishnomalar soni"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user_count = UserNotification.objects.filter(
            user=request.user,
            is_read=False
        ).count()
        
        app_count = ApplicationNotification.objects.filter(
            user=request.user,
            is_read=False
        ).count()
        
        return Response({
            "unread_count": user_count + app_count,
            "user_notifications": user_count,
            "application_notifications": app_count
        })


# ==================== ADMIN NOTIFICATION MANAGEMENT ====================
class NotificationCreateView(generics.CreateAPIView):
    """Admin tomonidan bildirishnoma yaratish"""
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    parser_classes = [MultiPartParser, FormParser, JSONParser]


class NotificationListView(generics.ListAPIView):
    """Admin tomonidan yaratilgan bildirishnomalar ro'yxati"""
    serializer_class = NotificationSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['target_type', 'is_active']
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Notification.objects.none()
        return Notification.objects.all().order_by('-created_at')


class NotificationDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Admin tomonidan yaratilgan bildirishnomani ko'rish/tahrirlash/o'chirish"""
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    parser_classes = [MultiPartParser, FormParser, JSONParser]



class AssignRoomToStudentView(generics.UpdateAPIView):
    """Talabaga qavat va xona biriktirish"""
    queryset = Student.objects.all()
    serializer_class = AssignRoomSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    lookup_field = 'id'
    lookup_url_kwarg = 'student_id'
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Student.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Student.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Student.objects.filter(dormitory__admin=user)
        return Student.objects.none()
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        return Response({
            "message": "Talaba muvaffaqiyatli xonaga joylashtirildi",
            "student": {
                "id": instance.id,
                "name": f"{instance.name} {instance.last_name}",
                "floor": instance.floor.name if instance.floor else None,
                "room": instance.room.name if instance.room else None,
                "is_active": instance.is_active,
                "placement_status": instance.placement_status
            }
        }, status=status.HTTP_200_OK)


class RemoveStudentFromRoomView(generics.UpdateAPIView):
    """Talabani xonadan chiqarish"""
    queryset = Student.objects.all()
    serializer_class = RemoveRoomSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    lookup_field = 'id'
    lookup_url_kwarg = 'student_id'
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Student.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Student.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Student.objects.filter(dormitory__admin=user)
        return Student.objects.none()
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        return Response({
            "message": "Talaba xonadan chiqarildi",
            "student": {
                "id": instance.id,
                "name": f"{instance.name} {instance.last_name}",
                "is_active": instance.is_active,
                "placement_status": instance.placement_status
            }
        }, status=status.HTTP_200_OK)


class UnassignedStudentsView(generics.ListAPIView):
    """Xonaga joylashtirilmagan talabalar ro'yxati"""
    serializer_class = StudentSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['name', 'last_name', 'passport']
    filterset_fields = ['dormitory', 'gender', 'course']
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Student.objects.none()
        
        user = self.request.user
        queryset = Student.objects.filter(is_active=False, room__isnull=True)
        
        if user.is_superuser:
            return queryset
        elif hasattr(user, 'role') and user.role == 'admin':
            return queryset.filter(dormitory__admin=user)
        
        return Student.objects.none()



# ==================== ADMIN DASHBOARD STATISTICS ====================
class AdminDashboardStatsView(APIView):
    """Admin uchun to'liq statistika - bitta endpoint"""
    permission_classes = [IsAdminOrDormitoryAdmin]

    def get(self, request):
        from django.db.models import Sum, Count, Q
        from django.utils import timezone
        from datetime import timedelta, date
        import calendar

        user = request.user

        if user.is_superuser:
            students = Student.objects.all()
            rooms = Room.objects.all()
            payments = Payment.objects.all()
            applications = Application.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            dormitory = Dormitory.objects.filter(admin=user).first()
            if not dormitory:
                return Response(
                    {"error": "Sizga biriktirilgan yotoqxona topilmadi"},
                    status=status.HTTP_404_NOT_FOUND
                )
            students = Student.objects.filter(dormitory=dormitory)
            rooms = Room.objects.filter(floor__dormitory=dormitory)
            payments = Payment.objects.filter(dormitory=dormitory)
            applications = Application.objects.filter(dormitory=dormitory)
        else:
            return Response({"error": "Ruxsat yo'q"}, status=status.HTTP_403_FORBIDDEN)

        # ── Talabalar ──────────────────────────────────────────────
        total_students  = students.count()
        male_students   = students.filter(gender='Erkak').count()
        female_students = students.filter(gender='Ayol').count()
        active_students   = students.filter(is_active=True).count()
        inactive_students = students.filter(is_active=False).count()

        students_by_course = {
            course: students.filter(course=course).count()
            for course in ['1-kurs', '2-kurs', '3-kurs', '4-kurs', '5-kurs']
        }

        # ── Xonalar (real Student soni bilan) ─────────────────────
        # current_occupancy ni DB dan emas, real Student queryset dan olamiz
        room_ids = list(rooms.values_list('id', flat=True))

        # Har bir xona uchun faol talabalar soni
        occupied_per_room = (
            Student.objects.filter(room_id__in=room_ids, is_active=True)
            .values('room_id')
            .annotate(cnt=Count('id'))
        )
        occupied_map = {r['room_id']: r['cnt'] for r in occupied_per_room}

        total_capacity = rooms.aggregate(s=Sum('capacity'))['s'] or 0
        total_occupied = sum(occupied_map.values())
        total_free     = max(total_capacity - total_occupied, 0)

        male_rooms   = rooms.filter(gender='male')
        female_rooms = rooms.filter(gender='female')

        male_room_ids   = list(male_rooms.values_list('id', flat=True))
        female_room_ids = list(female_rooms.values_list('id', flat=True))

        male_capacity   = male_rooms.aggregate(s=Sum('capacity'))['s'] or 0
        female_capacity = female_rooms.aggregate(s=Sum('capacity'))['s'] or 0

        male_occupied   = sum(v for k, v in occupied_map.items() if k in male_room_ids)
        female_occupied = sum(v for k, v in occupied_map.items() if k in female_room_ids)
        male_free   = max(male_capacity - male_occupied, 0)
        female_free = max(female_capacity - female_occupied, 0)

        def occ_rate(occ, cap):
            return round(occ / cap * 100, 1) if cap > 0 else 0.0

        # Status hisoblash (real ma'lumot asosida)
        available_rooms          = sum(1 for rid in room_ids if occupied_map.get(rid, 0) == 0)
        fully_occupied_rooms     = 0
        partially_occupied_rooms = 0
        for r in rooms:
            occ = occupied_map.get(r.id, 0)
            cap = r.capacity or 0
            if cap > 0 and occ >= cap:
                fully_occupied_rooms += 1
            elif occ > 0:
                partially_occupied_rooms += 1

        # ── To'lovlar ─────────────────────────────────────────────
        total_payments    = payments.count()
        approved_payments = payments.filter(status='APPROVED').count()
        cancelled_payments = payments.filter(status='CANCELLED').count()
        total_amount = payments.filter(status='APPROVED').aggregate(s=Sum('amount'))['s'] or 0

        thirty_days_ago = timezone.now() - timedelta(days=30)
        paid_student_ids = set(
            payments.filter(status='APPROVED', paid_date__gte=thirty_days_ago)
            .values_list('student_id', flat=True).distinct()
        )
        paid_students_count = len(paid_student_ids)
        # Faqat faol talabalar orasidan qarzdorlar
        active_student_ids = set(students.filter(is_active=True).values_list('id', flat=True))
        debtors_count = max(len(active_student_ids - paid_student_ids), 0)

        # ── Arizalar ──────────────────────────────────────────────
        total_applications     = applications.count()
        pending_applications   = applications.filter(status='Pending').count()
        approved_applications  = applications.filter(status='Approved').count()
        rejected_applications  = applications.filter(status='Rejected').count()
        cancelled_applications = applications.filter(status='Cancelled').count()

        # ── Daromad ───────────────────────────────────────────────
        now = timezone.now()

        monthly_income = payments.filter(
            status='APPROVED',
            paid_date__year=now.year,
            paid_date__month=now.month
        ).aggregate(s=Sum('amount'))['s'] or 0

        yearly_income = payments.filter(
            status='APPROVED',
            paid_date__year=now.year
        ).aggregate(s=Sum('amount'))['s'] or 0

        # Oxirgi 6 oy (to'g'ri kalendar hisobi)
        monthly_chart = []
        for i in range(5, -1, -1):
            # Joriy oydan i oy oldin
            month = now.month - i
            year  = now.year
            while month <= 0:
                month += 12
                year  -= 1
            m_income = payments.filter(
                status='APPROVED',
                paid_date__year=year,
                paid_date__month=month
            ).aggregate(s=Sum('amount'))['s'] or 0
            monthly_chart.append({
                "month": f"{year}-{month:02d}",
                "income": m_income
            })

        return Response({
            "students": {
                "total":    total_students,
                "male":     male_students,
                "female":   female_students,
                "active":   active_students,
                "inactive": inactive_students,
                "by_course": students_by_course,
            },
            "rooms": {
                "total": {
                    "rooms":          rooms.count(),
                    "capacity":       total_capacity,
                    "occupied":       total_occupied,
                    "free":           total_free,
                    "occupancy_rate": occ_rate(total_occupied, total_capacity),
                },
                "male": {
                    "rooms":          male_rooms.count(),
                    "capacity":       male_capacity,
                    "occupied":       male_occupied,
                    "free":           male_free,
                    "occupancy_rate": occ_rate(male_occupied, male_capacity),
                },
                "female": {
                    "rooms":          female_rooms.count(),
                    "capacity":       female_capacity,
                    "occupied":       female_occupied,
                    "free":           female_free,
                    "occupancy_rate": occ_rate(female_occupied, female_capacity),
                },
                "by_status": {
                    "available":          available_rooms,
                    "partially_occupied": partially_occupied_rooms,
                    "fully_occupied":     fully_occupied_rooms,
                },
            },
            "payments": {
                "total":        total_payments,
                "approved":     approved_payments,
                "cancelled":    cancelled_payments,
                "total_amount": total_amount,
                "paid_students": paid_students_count,
                "debtors":      debtors_count,
            },
            "applications": {
                "total":     total_applications,
                "pending":   pending_applications,
                "approved":  approved_applications,
                "rejected":  rejected_applications,
                "cancelled": cancelled_applications,
            },
            "income": {
                "monthly":       monthly_income,
                "yearly":        yearly_income,
                "monthly_chart": monthly_chart,
            },
        }, status=status.HTTP_200_OK)



class MyDormitoryView(generics.RetrieveUpdateAPIView):
    """Admin o'z yotoqxonasini ko'rish va tahrirlash"""
    serializer_class = DormitorySerializer
    permission_classes = [IsDormitoryAdmin]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_object(self):
        if getattr(self, 'swagger_fake_view', False):
            return None
        
        user = self.request.user
        
        # Superuser barcha yotoqxonalarni ko'radi (birinchisini)
        if user.is_superuser:
            dormitory = Dormitory.objects.first()
            if not dormitory:
                raise NotFound("Hech qanday yotoqxona topilmadi")
            return dormitory
        
        # Admin o'z yotoqxonasini ko'radi
        try:
            return Dormitory.objects.get(admin=user)
        except Dormitory.DoesNotExist:
            raise NotFound("Sizga biriktirilgan yotoqxona topilmadi")


class MyDormitoriesListView(generics.ListAPIView):
    """Admin o'z yotoqxonalari ro'yxati (agar bir nechta bo'lsa)"""
    serializer_class = DormitorySerializer
    permission_classes = [IsDormitoryAdmin]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['name', 'address']
    filterset_fields = ['is_active']
    
    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Dormitory.objects.none()
        
        user = self.request.user
        
        # Superuser barcha yotoqxonalarni ko'radi
        if user.is_superuser:
            return Dormitory.objects.all()
        
        # Admin o'z yotoqxonalarini ko'radi
        return Dormitory.objects.filter(admin=user)


class StudentApplicationView(generics.RetrieveAPIView):
    """Talabaning o'z arizasini ko'rish"""
    serializer_class = ApplicationSerializer
    permission_classes = [IsAuthenticated]
    
    def get_object(self):
        try:
            return Application.objects.get(user=self.request.user)
        except Application.DoesNotExist:
            raise NotFound("Sizning arizangiz topilmadi")


class GeneralStatsView(APIView):
    """Umumiy statistika - yotoqxonalar, foydalanuvchilar, universitetlar soni"""
    permission_classes = [AllowAny]  # Hamma ko'ra oladi

    def get(self, request):
        # Yotoqxonalar statistikasi
        total_dormitories = Dormitory.objects.filter(is_active=True).count()
        active_dormitories = Dormitory.objects.filter(is_active=True).count()

        # Foydalanuvchilar statistikasi
        total_users = User.objects.count()
        active_users = User.objects.filter(is_active=True).count()
        students_count = User.objects.filter(role='student').count()
        admins_count = User.objects.filter(role='admin').count()
        ijarachi_count = User.objects.filter(role='ijarachi').count()
        sardor_count = User.objects.filter(role='sardor').count()

        # Universitetlar statistikasi
        total_universities = University.objects.count()

        # Arizalar statistikasi
        total_applications = Application.objects.count()
        pending_applications = Application.objects.filter(status='Pending').count()
        approved_applications = Application.objects.filter(status='Approved').count()
        rejected_applications = Application.objects.filter(status='Rejected').count()

        # Xonalar statistikasi (real Student soni bilan)
        from django.db.models import Sum, Count

        total_rooms = Room.objects.count()
        total_capacity = Room.objects.aggregate(Sum('capacity'))['capacity__sum'] or 0
        total_occupied = Student.objects.filter(is_active=True, room__isnull=False).count()
        total_free = max(total_capacity - total_occupied, 0)

        # Viloyatlar statistikasi
        total_provinces = Province.objects.count()

        # Kvartiralar statistikasi
        total_apartments = Apartment.objects.filter(is_active=True).count()

        return Response({
            "dormitories": {
                "total": total_dormitories,
                "active": active_dormitories
            },
            "users": {
                "total": total_users,
                "active": active_users,
                "students": students_count,
                "admins": admins_count,
                "ijarachi": ijarachi_count,
                "sardor": sardor_count
            },
            "universities": {
                "total": total_universities
            },
            "applications": {
                "total": total_applications,
                "pending": pending_applications,
                "approved": approved_applications,
                "rejected": rejected_applications
            },
            "rooms": {
                "total": total_rooms,
                "capacity": total_capacity,
                "occupied": total_occupied,
                "free": total_free,
                "occupancy_rate": round((total_occupied / total_capacity * 100) if total_capacity > 0 else 0, 1)
            },
            "provinces": {
                "total": total_provinces
            },
            "apartments": {
                "total": total_apartments
            }
        }, status=status.HTTP_200_OK)


# ==================== COMPLAINT VIEWS ====================
class ComplaintListView(generics.ListCreateAPIView):
    """Talaba shikoyat yuborish va ko'rish"""
    serializer_class = ComplaintSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'category']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Complaint.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Complaint.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Complaint.objects.filter(dormitory__admin=user)
        # Talaba faqat o'z shikoyatlarini ko'radi
        try:
            student = Student.objects.get(user=user)
            return Complaint.objects.filter(student=student)
        except Student.DoesNotExist:
            return Complaint.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        try:
            student = Student.objects.get(user=user)
        except Student.DoesNotExist:
            raise PermissionDenied("Faqat talabalar shikoyat yuborishi mumkin")
        if not student.dormitory:
            raise PermissionDenied("Siz hech qaysi yotoqxonaga biriktirilmagansiz")
        serializer.save(student=student, dormitory=student.dormitory)


class ComplaintDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Shikoyatni ko'rish / admin javob berish"""
    serializer_class = ComplaintSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Complaint.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Complaint.objects.all()
        elif hasattr(user, 'role') and user.role == 'admin':
            return Complaint.objects.filter(dormitory__admin=user)
        try:
            student = Student.objects.get(user=user)
            return Complaint.objects.filter(student=student)
        except Student.DoesNotExist:
            return Complaint.objects.none()

    def update(self, request, *args, **kwargs):
        """Admin status va javob o'zgartira oladi"""
        user = request.user
        instance = self.get_object()
        # Talaba faqat o'qiy oladi (update qila olmaydi)
        if hasattr(user, 'role') and user.role == 'student':
            return Response(
                {"error": "Talaba shikoyatni o'zgartira olmaydi"},
                status=status.HTTP_403_FORBIDDEN
            )
        # Admin status va admin_response ni yangilaydi
        allowed_fields = ['status', 'admin_response']
        data = {k: v for k, v in request.data.items() if k in allowed_fields}
        serializer = self.get_serializer(instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        complaint = serializer.save()

        # Talabaga bildirishnoma
        if complaint.student.user:
            ApplicationNotification.objects.create(
                user=complaint.student.user,
                message=f"Shikoyatingizga javob berildi: {complaint.admin_response or ''} (Status: {complaint.get_status_display()})"
            )
        return Response(serializer.data)


# ==================== STAFF VIEWS ====================
class StaffListView(generics.ListCreateAPIView):
    serializer_class = StaffSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['position', 'is_active']
    search_fields = ['name', 'last_name', 'phone']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Staff.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Staff.objects.all()
        return Staff.objects.filter(dormitory__admin=user)

    def perform_create(self, serializer):
        user = self.request.user
        if user.is_superuser:
            serializer.save()
            return
        try:
            dormitory = Dormitory.objects.get(admin=user)
        except Dormitory.DoesNotExist:
            raise PermissionDenied("Sizga tegishli yotoqxona topilmadi")
        serializer.save(dormitory=dormitory)


class StaffDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = StaffSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Staff.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Staff.objects.all()
        return Staff.objects.filter(dormitory__admin=user)


# ==================== STAFF ATTENDANCE VIEWS ====================
class StaffAttendanceListView(generics.ListCreateAPIView):
    serializer_class = StaffAttendanceSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['staff', 'date', 'status']

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return StaffAttendance.objects.none()
        user = self.request.user
        if user.is_superuser:
            return StaffAttendance.objects.all()
        return StaffAttendance.objects.filter(staff__dormitory__admin=user)


class StaffAttendanceDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = StaffAttendanceSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return StaffAttendance.objects.none()
        user = self.request.user
        if user.is_superuser:
            return StaffAttendance.objects.all()
        return StaffAttendance.objects.filter(staff__dormitory__admin=user)


# ==================== TRANSFER ROOM VIEW ====================
class TransferStudentRoomView(generics.UpdateAPIView):
    """Talabani xonadan xonaga o'tkazish (to'la xonadan ham)"""
    queryset = Student.objects.all()
    serializer_class = TransferRoomSerializer
    permission_classes = [IsAdminOrDormitoryAdmin]
    lookup_field = 'id'
    lookup_url_kwarg = 'student_id'

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Student.objects.none()
        user = self.request.user
        if user.is_superuser:
            return Student.objects.all()
        return Student.objects.filter(dormitory__admin=user)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        student = serializer.save()
        return Response({
            "message": "Talaba muvaffaqiyatli yangi xonaga o'tkazildi",
            "student": {
                "id": student.id,
                "name": f"{student.name} {student.last_name}",
                "new_floor": student.floor.name if student.floor else None,
                "new_room": student.room.name if student.room else None,
            }
        }, status=status.HTTP_200_OK)


# ==================== FLOOR LEADER DASHBOARD ====================
class FloorLeaderDashboardView(APIView):
    """Sardor uchun to'liq dashboard"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        try:
            floor_leader = FloorLeader.objects.select_related('floor', 'floor__dormitory').get(user=user)
        except FloorLeader.DoesNotExist:
            return Response({"error": "Siz qavat sardori emassiz"}, status=status.HTTP_403_FORBIDDEN)

        floor = floor_leader.floor
        students = Student.objects.filter(floor=floor, is_active=True)

        # Davomat statistikasi (bugungi)
        from django.utils import timezone
        today = timezone.now().date()
        today_session = AttendanceSession.objects.filter(floor=floor, date=today).first()
        today_present = 0
        today_absent = 0
        if today_session:
            today_present = AttendanceRecord.objects.filter(session=today_session, status='in').count()
            today_absent = AttendanceRecord.objects.filter(session=today_session, status='out').count()

        # Oxirgi 7 kunlik davomat
        from datetime import timedelta
        last_7_days = []
        for i in range(6, -1, -1):
            day = today - timedelta(days=i)
            session = AttendanceSession.objects.filter(floor=floor, date=day).first()
            present = AttendanceRecord.objects.filter(session=session, status='in').count() if session else 0
            last_7_days.append({
                "date": str(day),
                "present": present,
                "total": students.count()
            })

        # Yig'im statistikasi
        from django.db.models import Sum
        collections = Collection.objects.filter(floor=floor)
        total_collections = collections.count()
        total_collected = CollectionRecord.objects.filter(
            collection__floor=floor, status="To'lagan"
        ).count()

        # Vazifalar
        tasks = TaskForLeader.objects.filter(user=user)

        return Response({
            "floor": {
                "id": floor.id,
                "name": floor.name,
                "gender": floor.gender,
                "dormitory": floor.dormitory.name,
            },
            "students": {
                "total": students.count(),
                "by_room": [
                    {
                        "room": r.name,
                        "capacity": r.capacity,
                        "occupied": Student.objects.filter(room=r, is_active=True).count(),
                        "free": max((r.capacity or 0) - Student.objects.filter(room=r, is_active=True).count(), 0),
                    }
                    for r in floor.room_set.all()
                ]
            },
            "attendance_today": {
                "has_session": today_session is not None,
                "session_id": today_session.id if today_session else None,
                "present": today_present,
                "absent": today_absent,
                "total": students.count(),
            },
            "attendance_last_7_days": last_7_days,
            "collections": {
                "total": total_collections,
                "paid_records": total_collected,
            },
            "tasks": {
                "total": tasks.count(),
                "pending": tasks.filter(status='PENDING').count(),
                "in_progress": tasks.filter(status='IN_PROGRESS').count(),
                "completed": tasks.filter(status='COMPLETED').count(),
            }
        }, status=status.HTTP_200_OK)


# ==================== ATTENDANCE FULL SESSION VIEW ====================
class AttendanceSessionFullCreateView(generics.CreateAPIView):
    """Sardor - davomat sessiyasi yaratish va talabalarni belgilash"""
    serializer_class = AttendanceSessionCreateSerializer
    permission_classes = [IsFloorLeader]

    def perform_create(self, serializer):
        user = self.request.user
        try:
            floor_leader = FloorLeader.objects.get(user=user)
        except FloorLeader.DoesNotExist:
            raise PermissionDenied("Siz qavat sardori emassiz")
        serializer.save(leader=floor_leader, floor=floor_leader.floor)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        session = serializer.instance
        records = AttendanceRecord.objects.filter(session=session).select_related('student')
        return Response({
            "message": "Davomat muvaffaqiyatli yaratildi",
            "session_id": session.id,
            "date": str(session.date),
            "floor": session.floor.name,
            "total_students": records.count(),
            "present": records.filter(status='in').count(),
            "absent": records.filter(status='out').count(),
            "records": [
                {
                    "student_id": r.student.id,
                    "student_name": f"{r.student.name} {r.student.last_name}",
                    "status": r.status
                }
                for r in records
            ]
        }, status=status.HTTP_201_CREATED)


class AttendanceRecordUpdateView(generics.UpdateAPIView):
    """Sardor - bitta talaba davomatini o'zgartirish"""
    queryset = AttendanceRecord.objects.all()
    serializer_class = AttendanceRecordSerializer
    permission_classes = [IsFloorLeader]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return AttendanceRecord.objects.none()
        user = self.request.user
        if user.is_superuser:
            return AttendanceRecord.objects.all()
        floor_leader = FloorLeader.objects.filter(user=user).first()
        if floor_leader:
            return AttendanceRecord.objects.filter(session__floor=floor_leader.floor)
        return AttendanceRecord.objects.none()


# ==================== EXCEL EXPORT VIEWS ====================
class ExportStudentsExcelView(APIView):
    """Talabalar ro'yxatini Excel ga export qilish"""
    permission_classes = [IsAdminOrDormitoryAdmin]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        user = request.user

        # Queryset
        if user.is_superuser:
            students = Student.objects.select_related(
                'dormitory', 'floor', 'room', 'province', 'district'
            ).all()
        elif hasattr(user, 'role') and user.role == 'admin':
            dormitory = Dormitory.objects.filter(admin=user).first()
            if not dormitory:
                return Response({"error": "Yotoqxona topilmadi"}, status=404)
            students = Student.objects.select_related(
                'dormitory', 'floor', 'room', 'province', 'district'
            ).filter(dormitory=dormitory)
        else:
            return Response({"error": "Ruxsat yo'q"}, status=403)

        # Filter parametrlari
        floor_id = request.query_params.get('floor')
        room_id = request.query_params.get('room')
        gender = request.query_params.get('gender')
        course = request.query_params.get('course')
        is_active = request.query_params.get('is_active')

        if floor_id:
            students = students.filter(floor_id=floor_id)
        if room_id:
            students = students.filter(room_id=room_id)
        if gender:
            students = students.filter(gender=gender)
        if course:
            students = students.filter(course=course)
        if is_active is not None:
            students = students.filter(is_active=is_active.lower() == 'true')

        # Excel yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Talabalar"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style='thin', color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "№", "Ism", "Familiya", "Otasining ismi", "Jinsi", "Kurs",
            "Fakultet", "Guruh", "Telefon", "Pasport",
            "Viloyat", "Tuman", "Yotoqxona", "Qavat", "Xona",
            "Holati", "Joylashuv holati", "Faolmi"
        ]

        col_widths = [4, 15, 15, 15, 8, 8, 20, 12, 15, 12, 15, 15, 20, 12, 10, 15, 18, 8]

        for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 30

        # Ma'lumotlar
        row_fill_even = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
        data_align = Alignment(horizontal="left", vertical="center")

        for idx, student in enumerate(students, 1):
            row = idx + 1
            fill = row_fill_even if idx % 2 == 0 else None
            values = [
                idx,
                student.name or '',
                student.last_name or '',
                student.middle_name or '',
                student.gender or '',
                student.course or '',
                student.faculty or '',
                student.group or '',
                student.phone or '',
                student.passport or '',
                student.province.name if student.province else '',
                student.district.name if student.district else '',
                student.dormitory.name if student.dormitory else '',
                student.floor.name if student.floor else '',
                student.room.name if student.room else '',
                student.status or '',
                student.placement_status or '',
                'Ha' if student.is_active else "Yo'q",
            ]
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_num, value=value)
                cell.alignment = data_align
                cell.border = border
                if fill:
                    cell.fill = fill

        ws.freeze_panes = "A2"

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="talabalar.xlsx"'
        wb.save(response)
        return response


class ExportPaymentsExcelView(APIView):
    """To'lovlar jadvalini Excel ga export qilish"""
    permission_classes = [IsAdminOrDormitoryAdmin]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        user = request.user

        if user.is_superuser:
            payments = Payment.objects.select_related('student', 'dormitory').all()
        elif hasattr(user, 'role') and user.role == 'admin':
            dormitory = Dormitory.objects.filter(admin=user).first()
            if not dormitory:
                return Response({"error": "Yotoqxona topilmadi"}, status=404)
            payments = Payment.objects.select_related('student', 'dormitory').filter(dormitory=dormitory)
        else:
            return Response({"error": "Ruxsat yo'q"}, status=403)

        # Filter
        student_id = request.query_params.get('student')
        pay_status = request.query_params.get('status')
        method = request.query_params.get('method')
        year = request.query_params.get('year')
        month = request.query_params.get('month')

        if student_id:
            payments = payments.filter(student_id=student_id)
        if pay_status:
            payments = payments.filter(status=pay_status)
        if method:
            payments = payments.filter(method=method)
        if year:
            payments = payments.filter(paid_date__year=year)
        if month:
            payments = payments.filter(paid_date__month=month)

        payments = payments.order_by('-paid_date')

        # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "To'lovlar"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style='thin', color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "№", "Talaba ismi", "Familiya", "Xona", "Qavat",
            "Yotoqxona", "Summa (so'm)", "To'lov sanasi",
            "Amal qilish muddati", "Usul", "Holati", "Izoh"
        ]
        col_widths = [4, 15, 15, 10, 10, 20, 15, 18, 18, 10, 12, 25]

        for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 30

        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        data_align = Alignment(horizontal="left", vertical="center")

        total_amount = 0
        for idx, payment in enumerate(payments, 1):
            row = idx + 1
            paid_date = payment.paid_date.strftime('%Y-%m-%d %H:%M') if payment.paid_date else ''
            valid_until = str(payment.valid_until) if payment.valid_until else ''
            amount = payment.amount or 0
            if payment.status == 'APPROVED':
                total_amount += amount

            values = [
                idx,
                payment.student.name if payment.student else '',
                payment.student.last_name if payment.student else '',
                payment.student.room.name if payment.student and payment.student.room else '',
                payment.student.floor.name if payment.student and payment.student.floor else '',
                payment.dormitory.name if payment.dormitory else '',
                amount,
                paid_date,
                valid_until,
                payment.method or '',
                payment.status or '',
                payment.comment or '',
            ]
            row_fill = green_fill if payment.status == 'APPROVED' else red_fill
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_num, value=value)
                cell.alignment = data_align
                cell.border = border
                cell.fill = row_fill

        # Jami qator
        total_row = payments.count() + 2
        ws.cell(row=total_row, column=6, value="JAMI:").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=7, value=total_amount)
        total_cell.font = Font(bold=True, color="375623")

        ws.freeze_panes = "A2"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="tolovlar.xlsx"'
        wb.save(response)
        return response
